from fastapi import APIRouter, Depends, HTTPException
from sqlmodel import Session, select
from typing import List
from sqlalchemy.orm import selectinload
from model.models import Ticket,User
from core.db import get_session
from utils import generate_ticket_id
from core.auth import get_current_user
from schema import TicketCreate,TicketPublic,TicketUpdate
from datetime import date
from model.models import Category,PriorityEnum,StatusEnum
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/tickets", tags=["Tickets"])


@router.get("/export")
def export_tickets(session: Session = Depends(get_session),current_user: User = Depends(get_current_user),q:str|None=None):
    if q:
        tickets = session.exec(select(Ticket).where(Ticket.assigned_to == current_user)).all()
    else:
        tickets = session.exec(select(Ticket)).all()
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Incident & Ticket Log"

    # ------------------ STYLES ------------------
    header_fill = PatternFill("solid", fgColor="00E5E5")
    title_fill = PatternFill("solid", fgColor="FFF2CC")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    center = Alignment(horizontal="center", vertical="center")

    status_colors = {
        "Closed": PatternFill("solid", fgColor="92D050"),
        "Open": PatternFill("solid", fgColor="FF6666"),
        "In Progress": PatternFill("solid", fgColor="FFD966"),
        "Pending": PatternFill("solid", fgColor="FFD966"),
    }

    # ------------------ TITLE ------------------
    ws.merge_cells("A1:I1")
    ws["A1"] = "ACPL - Incident & Ticket Log"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws["A1"].fill = title_fill

    # ------------------ HEADERS ------------------
    headers = [
        "Sl.no",
        "Reported By",
        "Date Reported",
        "Issue Description",
        "Assigned To",
        "Priority (Low/Medium/High)",
        "Status (Open/In Progress/Closed)",
        "Root Cause",
        "Resolution Summary",
        "Date Closed"
    ]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # ------------------ DATA ------------------
    row_no = 3
    for idx, t in enumerate(tickets, start=1):
        ws.append([
            idx,
            t.reported_by.username if t.reported_by else "",
            t.date_reported.strftime("%d-%m-%Y") if t.date_reported else "",
            t.issue_description,
            t.assigned_to.username if t.assigned_to else "",
            t.priority.value if t.priority else "",
            t.status.value if t.status else "",
            t.root_cause or "",
            t.resolution_summary or "",
            t.date_closed or "",
        ])

        for col in range(1, 11):
            cell = ws.cell(row=row_no, column=col)
            cell.border = border

            # Status coloring
            if col == 7:
                fill = status_colors.get(cell.value)
                if fill:
                    cell.fill = fill
                    cell.alignment = center

        row_no += 1

    # ------------------ AUTO WIDTH ------------------
    for i, col in enumerate(ws.iter_cols(min_row=2), start=1):
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(i)].width = max_len + 3

    # ------------------ STREAM ------------------
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=Tickets.xlsx"
        },
    )

@router.get("/assigned/me", response_model=List[TicketPublic])
def my_assigned_tickets(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    stmt = select(Ticket).where(
        Ticket.assigned_to_id == current_user.id
    )
    tickets = session.exec(stmt.options(
            selectinload(Ticket.reported_by),
            selectinload(Ticket.assigned_to),
        )).all()
    result = []

    for t in tickets:
        result.append(
            TicketPublic(
                id=t.id,
                ticket_id=t.ticket_id,
                category=t.category,
                issue_description=t.issue_description,
                priority=t.priority,
                status=t.status,
                date_reported=t.date_reported,
                date_closed=t.date_closed,
                root_cause=t.root_cause,
                resolution_summary=t.resolution_summary,
                reported_by=t.reported_by.username,
                assigned_to=t.assigned_to.username if t.assigned_to else None,
            )
        )

    return result

    
@router.post("/", response_model=Ticket)
def create_ticket(
    payload: TicketCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user)
):
    ticket = Ticket(
        ticket_id=generate_ticket_id(session),
        reported_by_id=current_user.id,
        category=payload.category,
        issue_description=payload.issue_description,
        priority=payload.priority,
        assigned_to_id = payload.assigned_to_id
    )

    session.add(ticket)
    session.commit()
    session.refresh(ticket)

    return ticket

@router.get("/", response_model=List[TicketPublic])
def list_tickets(
    status: StatusEnum | None = None,
    priority: PriorityEnum | None = None,
    session: Session = Depends(get_session)
):
    stmt = select(Ticket)

    if status:
        stmt = stmt.where(Ticket.status == status)
    if priority:
        stmt = stmt.where(Ticket.priority == priority)
    tickets = session.exec(stmt.options(
            selectinload(Ticket.reported_by),
            selectinload(Ticket.assigned_to),
        )).all()
    result = []

    for t in tickets:
        result.append(
            TicketPublic(
                id=t.id,
                ticket_id=t.ticket_id,
                category=t.category,
                issue_description=t.issue_description,
                priority=t.priority,
                status=t.status,
                date_reported=t.date_reported,
                date_closed=t.date_closed,
                root_cause=t.root_cause,
                resolution_summary=t.resolution_summary,
                reported_by=t.reported_by.username,
                assigned_to=t.assigned_to.username if t.assigned_to else None,
            )
        )

    return result


@router.get("/{ticket_id}", response_model=Ticket)
def get_ticket(
    ticket_id: int,
    session: Session = Depends(get_session)
):
    ticket = session.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return ticket


@router.put("/{id}", response_model=TicketUpdate)
def update_ticket(
    id: int,
    updated_ticket: TicketUpdate,
    session: Session = Depends(get_session)
):
    ticket = session.get(Ticket, id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    ticket_data = updated_ticket.model_dump(exclude_unset=True)

    # Handle status logic separately
    if "status" in ticket_data:
        if ticket_data["status"] == StatusEnum.closed:
            ticket.date_closed = ticket_data.get(
                "date_closed", date.today()
            )
        else:
            ticket.date_closed = None

    # Update remaining fields
    for key, value in ticket_data.items():
        setattr(ticket, key, value)

    session.commit()
    session.refresh(ticket)
    return ticket


# @router.post("/{ticket_id}/close", response_model=Ticket)
# def close_ticket(
#     ticket_id: int,
#     session: Session = Depends(get_session)
# ):
#     ticket = session.get(Ticket, ticket_id)
#     if not ticket:
#         raise HTTPException(status_code=404, detail="Ticket not found")

#     ticket.status = StatusEnum.closed
#     ticket.date_closed = date.today()

#     session.commit()
#     session.refresh(ticket)
#     return ticket
# @router.put("/{ticket_id}",response_model=TicketPublic)
# async def updateTicket(
#     ticket_id: int,
#     ticketIUpdate:TicketUpdate,
#     session: Session = Depends(get_session)
# ):
#     ticket = session.get(Ticket, ticket_id)
#     if not ticket:
#         raise HTTPException(status_code=404, detail="Ticket not found")
#     ticket_data = ticketIUpdate.model_dump()
#     for key, value in ticket_data.items():
#         setattr(ticket, key, value)

#     session.add(ticket)
#     session.commit()
#     session.refresh(ticket)
#     return ticket
